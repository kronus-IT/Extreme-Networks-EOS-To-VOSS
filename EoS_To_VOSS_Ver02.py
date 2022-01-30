##########################################################
# EOS_To_VOSS: Convert Extreme EOS to Extreme VOSS       #
# Written by Martin FLammia, Director, Kronus IT         #
#########################################################

from openpyxl import Workbook, load_workbook
import re
import os
import inquirer

def read_excel (xlsx_workbook,xlsx_worksheet_name):
    ##Need to get list of number of rows
    ##Then iterate throgh each row using the number as a key
    ##Then add all the row of information as key = colomn & value = row
    xlsx_headers = {}
    xlsx_value = {}
    xlsx_rows = {}
    #xlsx_config_row = {}
    xlsx_sentence = {}
    xlsx_key_counter = 0
    xlsx_sentence_counter = 0
    xlsx_row_counter = 0
    xlsx_worksheet = xlsx_workbook[xlsx_worksheet_name]
    #xlsx_number_of_rows = xlsx_worksheet.max_row
    #xlsx_number_of_columns = xlsx_worksheet.max_column
    for xlsx_row_keys in xlsx_worksheet.iter_rows(max_row=1,values_only=True):
        for xlsx_row_key in xlsx_row_keys:
            xlsx_key_counter += 1
            xlsx_headers[xlsx_key_counter] = xlsx_row_key
    for xlsx_row_values in xlsx_worksheet.iter_rows(min_row=2,values_only=True):
        xlsx_row_counter += 1
        xlsx_value_counter = 0
        for xlsx_row_value in xlsx_row_values:
            xlsx_value_counter += 1
            xlsx_value[xlsx_value_counter] = xlsx_row_value
        xlsx_rows[xlsx_row_counter] = xlsx_value
        xlsx_value = {}
    xlsx_paragraph = {}
    for xlsx_row in xlsx_rows.values():
        xlsx_sentence_counter += 1
        for xlsx_row_key in xlsx_row.keys():
            if xlsx_row_key in xlsx_headers.keys():
                xlsx_sentence[xlsx_headers[xlsx_row_key]] = str(xlsx_row[xlsx_row_key])
                xlsx_paragraph[xlsx_sentence_counter] = xlsx_sentence
        xlsx_sentence = {}
    return xlsx_paragraph

def create_excel(eos_config_parameters):
    list_of_paragraph_keys = list(eos_config_parameters.keys())
    total_keys = len(list_of_paragraph_keys)
    paragraph_values = []
    voss_config = {}
    voss_config_block = {}
    unique_column_headers = set()
    #Workbook is dynamically created by just calling the import workbook class
    wb = Workbook()
    #Workbook name
    dest_filename = 'configuration.xlsx'
    #Worbook is always created with ast least one worksheet. This will get the active worksheet
    ws = wb.active
    #Worksheets will be labelled by default sheet, sheet1, sheet2. This will change the title
    ws.title = "Configuraton"
    #This itereates through each of the keys listed in teh list_of_paragraph_keys
    for paragraph_key in list_of_paragraph_keys:
        #This will grap a single config paragraph based on the paragraph key
        paragraph = eos_config_parameters[paragraph_key]
        #This itereates through each of the keys
        for key in paragraph.keys():
            #This adds each of the keys to a set. As sets can only have unique values
            #Only the unqiue headers are recodered and later add as columns tiles.l
            unique_column_headers.add(key)
    ws.append(list(unique_column_headers))
    #This itereates through each of the keys listed in teh list_of_paragraph_keys
    for paragraph_key in list_of_paragraph_keys:
        #This will grap a single config paragraph based on the paragraph key
        paragraph = eos_config_parameters[paragraph_key]
        for column_header in list(unique_column_headers):
            paragraph_values.append(paragraph[column_header])
            #paragraph_values = list(paragraph.values())
        #This now iterated through each line in paragraph extracting the data.
        ws.append (paragraph_values)
        paragraph_values = []
    #This saves the workbook
    wb.save(dest_filename)
    return unique_column_headers

def eos_paragraph_splitter (eos_config):
    '''This function takes a complete EOS interface configuration and iterates for each line, breaking the config into
    paragrpahs of code that start with 'interface ' and end in ' exit'. This so the data like IP address, VLAN ID, VIP
    etc can be extracted on a per interface basis'''
    #This setups the number of interface counts as an integer varible
    interface_count = 0
    #This initialises the paragraph count to 0, then ingrements through each iteration
    paragraph_count = 0
    #This stores each paragraph against a numbered key in a dictionary
    config_paragraph = {}
    #This holds the configs per line as the codee interates through the paragraph.
    #This is then what is added to the conf_paragraph dictionary as the
    #completed per paragraph set of config lines.
    config_lines = []
    #This funtion iterates through the config, splits it into each line, and strips the leading whitespace
    ###eos_config_per_line_no_leading_whitespace = eos_remove_leading_whitespace(eos_config)
    #This counts how many times "interface " appears in config as a means to count number of paragraphs
    #Needed to put a space after the interace as the command "interface-up-delay" was counting interface an extra time
    ##interface_count = eos_config.count("interface vlan.0.10")
    interface_count = sum('interface vlan' in s for s in eos_config)
    #This generates the loop that iterates through the test the number of times that is equal to the interace count
    while paragraph_count < interface_count:
        #This iterates through each line of code
        for first_loop in eos_config:
            #This appends each line of code to the "config_lines" varible
            config_lines.append(first_loop)
            #This looks out for any like of code that has " exit" in it
            if first_loop == "exit":
                #This appends all the conig_lines currently gathered and adds it to a dictionary
                #Were the key os the paragraph count
                config_paragraph[paragraph_count] = config_lines
                #This wipes all the config_lines, so as to process the next loop through the config
                config_lines = []
                #This keeps appending the paragraph count until it matches the interface_count and breaks the while loop
                paragraph_count += 1
    return config_paragraph


def eos_remove_leading_whitespace(eos_config):
    eos_config_per_line_no_leading_whitespace = []
    eos_config_per_line_left_strip = []
    eos_config_per_line_leading_whitespace = eos_config.split("\n")
    for eos_config_per_line in eos_config_per_line_leading_whitespace:
        eos_config_per_line_left_strip = eos_config_per_line.lstrip()
        eos_config_per_line_no_leading_whitespace.append(eos_config_per_line_left_strip)
    return eos_config_per_line_no_leading_whitespace

def eos_interface_grabber(full_eos_config):
    eos_interface_config = []
    start_capture = 0
    eos_full_config_clean = eos_remove_leading_whitespace(full_eos_config)
    for config_line in eos_full_config_clean:
        if start_capture >= 1:
            eos_interface_config.append(config_line)
            if config_line == "!" or "interface tun" in config_line:
                start_capture = 0
                return eos_interface_config
                exit()
        if config_line.startswith("interface vlan"):
            start_capture += 1
            if start_capture == 1:
                eos_interface_config.append(config_line)

def eos_vlan_grabber(full_eos_config):
    eos_vlan_config = []
    eos_vlans = {}
    start_capture = 0
    eos_full_config_clean = eos_remove_leading_whitespace(full_eos_config)
    for config_line in eos_full_config_clean:
        if start_capture >= 1:
            eos_vlan_config.append(config_line)
            eos_vlan_dict(eos_vlan_config)
            eos_vlans.update(eos_vlan_dict(eos_vlan_config))
            eos_vlan_config = []
            if config_line == "clear vlan egress" or "set vlan egress" in config_line:
                start_capture = 0
                return eos_vlans
                exit()
        if config_line.startswith("set vlan name"):
            start_capture += 1
            if start_capture == 1:
                eos_vlan_config.append(config_line)
                eos_vlans.update(eos_vlan_dict(eos_vlan_config))
                eos_vlan_config = []
    return

def eos_vlan_dict(eos_vlan_config):
    eos_vlan_config_string = str(eos_vlan_config)
    eos_vlan_id_vlan_name = {}
    eos_vlan_lines = eos_vlan_config[0].split(" ")
    #This grabs the third item in the list created from the split, which is the VLAN ID
    eos_vlan_id = eos_vlan_lines[3]
    #This grabs all text between speach marks
    eos_vlan_name = eos_vlan_lines[4:]
    eos_vlan_name = ' '.join(eos_vlan_name)
    eos_vlan_name_stripped = re.sub(r"[^a-zA-Z0-9-_ ]","",eos_vlan_name)
    ##eos_vlan_name = re.findall('"([^"]*)"',eos_vlan_config_string)
    #This formats the string to be vlan_10 so as to use as key for dictionary
    ##eos_vlan_id_concantenated = "vlan_" + eos_vlan_id
    eos_vlan_id_concantenated = eos_vlan_id
    eos_vlan_id_vlan_name[eos_vlan_id_concantenated] = eos_vlan_name_stripped
    return eos_vlan_id_vlan_name

def eos_config_extractor(config_paragraph):
    '''This function iterates through each paragraph of EOS configuration that has been presneted as a
    dictionary using the the eos_parapgraph_splitter. This provides a key / value pair for each instance
    of interface configuraton that this function then itterates through collect all the specific information 
    like ip address, VLAN ID, VIP, IP helper etc'''
    #This producces a list that contains a list of key that correspondes to each interface paragraph.
    list_of_paragraph_keys = list(config_paragraph.keys())
    eos_config_instances = {}
    eos_config_parameters = {}
    helper_ip_1 = ['']
    helper_ip_2 = ['']
    helper_ip_3 = ['']
    helper_ip_4 = ['']
    i_sid = '1211'
    vrrp_ip = ''
    #This interates through the paragraph keys, and each itteration puts the paragraph into the paragraph varible.
    for paragraph_key in list_of_paragraph_keys:
        #This grabs the paragraph based on the paragraph key and stores it in paragraph
        #This will essentially loop through each parapgrah grabing one at a time and storing it paragraph
        paragraph = config_paragraph[paragraph_key]
        #This now iterated through each line in paragraph extracting the data.
        for config_line in paragraph:
            if config_line.startswith("interface"):
                if config_line.count('loop') == 1:
                    print ("Please remove any configuration that has an 'interface loop.x.x' in the configuration file")
                    exit()
                if config_line.count('tun') == 1:
                    print ("Please remove any configuration that has an 'interface tun.x.x' in the configuration file")
                    exit()
                vlan_id = config_line.split("interface vlan.0.",1)[1]
                if (int(vlan_id) % 2) == 0:
                    #This is return if the VLAN ID is even
                    vrid = str(112)
                    vrrp_priority = str(200)
                if (int(vlan_id) % 2) != 0:
                    #This is return if the VLAN ID is even
                    vrid = str(111)
                    vrrp_priority = str(150)
            if config_line.startswith("vrrp address"):
                vrrp_ip = config_line.split(" ",3)[3]
            if config_line.startswith("ip helper-address"):
                if len(helper_ip_1[0]) == 0:
                    helper_ip_1[0] = config_line.split("ip helper-address ",1)[1]
                elif len(helper_ip_2[0]) == 0:
                    helper_ip_2[0] = config_line.split("ip helper-address ",1)[1]
                elif len(helper_ip_3[0]) == 0:
                    helper_ip_3[0] = config_line.split("ip helper-address ",1)[1]
                elif len(helper_ip_4[0]) == 0:
                    helper_ip_4[0] = config_line.split("ip helper-address ",1)[1]
            if config_line.startswith("ip address"):
                ip_address = config_line.split(" ")[2]
                ip_subnet = config_line.split(" ")[3]
        #This takes all the varibles and creates a dictionary of the varible name against the value
        eos_config_parameters.update({'vlan_id':vlan_id,'i_sid':i_sid,'vrrp_ip':vrrp_ip,'helper_ip_1':helper_ip_1[0],'helper_ip_2':helper_ip_2[0],'helper_ip_3':helper_ip_3[0],'helper_ip_4':helper_ip_4[0],'ip_address':ip_address,'ip_subnet':ip_subnet,'vrid':vrid, 'vrrp_priority':vrrp_priority})
        eos_config_instances[paragraph_key] = eos_config_parameters
        eos_config_parameters = {}
        helper_ip_1 = ['']
        helper_ip_2 = ['']
        helper_ip_3 = ['']
        helper_ip_4 = ['']
        vrrp_ip = ''
    return eos_config_instances

def config_param(eos_config_parameters,voss_include_vlan_create,voss_isid,eos_vlan_config):
    list_of_paragraph_keys = list(eos_config_parameters.keys())
    area = '0.0.0.0'
    vrrp_version = '3'
    voss_config = {}
    voss_config_block = {}
    voss_config_instances = {}
    full_voss_config =[]
    for paragraph_key in list_of_paragraph_keys:
        paragraph = eos_config_parameters[paragraph_key]
        if voss_include_vlan_create == 'y':
            #The try first checks a dictonary value exists against the key name vlan_name
            #The key with vlan_name is only present when using the spreadsheet import
            #So the try will skip past the try and into except, which is used with the text import
            try:
                #This checks the a vlaue is present againt the the key called vlan_name, then creates VLAN with the name. If not, jumps to except.
                if paragraph["vlan_name"]: 
                    voss_config['vlan_create'] = "vlan create " + paragraph["vlan_id"] + " name " + "\"" + paragraph["vlan_name"] + "\"" + " type port-mstprstp 0"
            except:
                #This just creates the vlan, no name.
                voss_config['vlan_create'] = "vlan create " + paragraph["vlan_id"] + " type port-mstprstp 0"
                #This checks if there is a vlan ID value againsts the vlan_id key in paragraph, has a matching vlan ID as a key againsts eos_vlan_config
                #This is because eos_vlan_config stors vlan ID's against VLAN names, scrapped from configurationl
                if paragraph["vlan_id"] in eos_vlan_config.keys():
                    #If there is a match, then the vlan name is used that is against the VLAN ID in eos_vlan_config
                    voss_config['vlan_name'] = "vlan " + paragraph["vlan_id"] + " name " + "\"" + eos_vlan_config[paragraph["vlan_id"]] + "\""
                else:
                    #If not, then the VLAN name becomes the vlan ID
                    voss_config['vlan_name'] = "vlan " + paragraph["vlan_id"] + " name " + "VLAN_" + paragraph["vlan_id"]
                i_sid_formula = paragraph["vlan_id"].rjust(4, "0")
                voss_config['isid_create'] = "vlan i-sid " + paragraph["vlan_id"] + " " + voss_isid + i_sid_formula
        voss_config['interface'] = "interface vlan " + paragraph["vlan_id"]
        voss_config['ip_address'] = "ip address " + paragraph["ip_address"] + " " + paragraph["ip_subnet"]
        voss_config['ip_vrrp_version'] = "ip vrrp version " + vrrp_version
        voss_config['ip_vrrp_address'] = "ip vrrp address " + paragraph["vrid"] + " " + paragraph["vrrp_ip"]
        voss_config['ip_vrrp_backup'] = "ip vrrp " + paragraph["vrid"] + " " + "backup-master enable"
        voss_config['ip_vrrp_enable'] = "ip vrrp " + paragraph["vrid"] + " " + "enable"
        voss_config['ip_vrrp_priority'] = "ip vrrp " + paragraph["vrid"] + " " + "priority " + paragraph["vrrp_priority"]
        voss_config['ip_dhcp_relay'] = "ip dhcp-relay"
        voss_config['ip_dhcp_relay_mode'] = "ip dhcp-relay mode dhcp"
        #voss_config['ip_ospf_area'] = "ip ospf area " + area
        voss_config['ip_ospf_passive'] = "ip ospf network passive"
        voss_config['ip_ospf_enable'] = "ip ospf enable"
        voss_config['ip_igmp_snooping'] = "ip igmp snooping"
        voss_config['config_exit'] = "exit"
        if paragraph["helper_ip_1"] != "":
            voss_config['ip_dhcp_1_ip'] = "ip dhcp-relay fwd-path " + paragraph["ip_address"] + " " + paragraph["helper_ip_1"]
            voss_config['ip_dhcp_1_enable'] = "ip dhcp-relay fwd-path " + paragraph["ip_address"] + " " + paragraph["helper_ip_1"] + " enable"
            voss_config['ip_dhcp_1_mode'] = "ip dhcp-relay fwd-path " + paragraph["ip_address"] + " " + paragraph["helper_ip_1"] + " mode bootp_dhcp"
        if paragraph["helper_ip_2"] != "":
            voss_config['ip_dhcp_2_ip'] = "ip dhcp-relay fwd-path " + paragraph["ip_address"] + " " + paragraph["helper_ip_2"]
            voss_config['ip_dhcp_2_enable'] = "ip dhcp-relay fwd-path " + paragraph["ip_address"] + " " + paragraph["helper_ip_2"] + " enable"
            voss_config['ip_dhcp_2_mode'] = "ip dhcp-relay fwd-path " + paragraph["ip_address"] + " " + paragraph["helper_ip_2"] + " mode bootp_dhcp"
        if paragraph["helper_ip_3"] != "":
            voss_config['ip_dhcp_3_ip'] = "ip dhcp-relay fwd-path " + paragraph["ip_address"] + " " + paragraph["helper_ip_3"]
            voss_config['ip_dhcp_3_enable'] = "ip dhcp-relay fwd-path " + paragraph["ip_address"] + " " + paragraph["helper_ip_3"] + " enable"
            voss_config['ip_dhcp_3_mode'] = "ip dhcp-relay fwd-path " + paragraph["ip_address"] + " " + paragraph["helper_ip_3"] + " mode bootp_dhcp"
        if paragraph["helper_ip_4"] != "":
            voss_config['ip_dhcp_4_ip'] = "ip dhcp-relay fwd-path " + paragraph["ip_address"] + " " + paragraph["helper_ip_4"]
            voss_config['ip_dhcp_4_enable'] = "ip dhcp-relay fwd-path " + paragraph["ip_address"] + " " + paragraph["helper_ip_4"] + " enable"
            voss_config['ip_dhcp_4_mode'] = "ip dhcp-relay fwd-path " + paragraph["ip_address"] + " " + paragraph["helper_ip_4"] + " mode bootp_dhcp"
        if paragraph["helper_ip_1"] != '':
            voss_config['exit'] = "exit"
        voss_config['newline'] = "\n"
        voss_config_block[paragraph_key] = voss_config
        voss_config = {}
    for key, value in voss_config_block.items():
        voss_config_output = value
        for key, value in voss_config_output.items():
            full_voss_config.append(value)
    return full_voss_config


introduction = '''This script produces VOSS VLAN / Interface configuration either based on values in an EXCEL spreedsheet or via EOS configuration in the form a text config back file.\n
The EOS backup file and / or the Excel (.xlsx) will need to be placed in the same location as this script file\n
A series of questions will be asked below that will produce the config based on what type of file you have present\n
Equally have have the option to produce an Excel spreedshet that list all the configuration details extracted like IP address, Subnet, VLAN ID etc\n
This program will extrapolate the following from an EOS back .txt file:

set vlan name 64 "Power House GUEST LAN"
interface vlan.0.64
  ip address 10.119.64.3 255.255.255.0 primary
  vrrp create 1 v2-IPv4
  vrrp address 1 110.119.64.1 
  vrrp priority 1 200
  vrrp accept-mode 1
  vrrp enable 1
  ip helper-address 10.119.0.140 
  ip helper-address 10.119.0.200 
  no ip redirects
  exit

And convert to something like the below. \n
* Even VLANs are given a VRID of 112 and priority of 200, odd VLANs are given a VRID of 111 and a priority of 150.
* If the VLAN name id not present the name will be VLAN_<VLAN ID>
* All interfaces are OSPF enable and passive.
* VRRP version 3 is used.

vlan create 64 name Power House GUEST LAN type port-mstprstp 0
vlan i-sid 64 12110064
interface vlan 64
ip address 10.119.255.10 255.255.255.0
ip vrrp version 3
ip vrrp address 2 112 
ip vrrp 112 backup-master enable
ip vrrp 112 enable
ip vrrp 112 priority 200
ip dhcp-relay
ip dhcp-relay mode dhcp
ip ospf area 0.0.0.0
ip ospf network passive
ip ospf enable
ip igmp snooping
exit
ip dhcp-relay fwd-path 10.119.255.10 10.119.0.140
ip dhcp-relay fwd-path 10.119.255.10 10.119.0.140 enable
ip dhcp-relay fwd-path 10.119.255.10 10.119.0.140 mode bootp_dhcp
ip dhcp-relay fwd-path 10.119.255.10 10.119.0.200 
ip dhcp-relay fwd-path 10.119.255.10 10.119.0.200 enable
ip dhcp-relay fwd-path 10.119.255.10 10.119.0.200 mode bootp_dhcp

exit
'''

print (introduction)
eos_text_confirm = str(input('Are you using a EOS back file in the form of a .txt file that you need to convert to VOSS, answer "y" or "n" : ') or "n").lower().strip()
eos_excel_confirm = str(input('Are you using the provided EXCEL template to convert into VOSS, answer "y" or "n": ') or "n").lower().strip()
produce_excel = str(input('Would you like an Excel spredsheet ceated that lists all the confiugraton details like IP Address, Subnet, VLAN ID; answer "y" or "n": ') or "n").lower().strip()
voss_include_vlan_create = str(input('Would like to include the VLAN creation with the interface configuration, answer "y" or "n": ') or "n").lower().strip()
voss_isid = str(input("What i-sid prefix will you be adding to you VLAN ID's ie. 1211xxxx. If not sure just hit return: ") or "1211").lower().strip()
if eos_text_confirm[0] == 'y':
    with open('eos_full_config.txt') as f:
        eos_full_config = f.read()
        f.close()
    eos_config_file = eos_interface_grabber(eos_full_config)
    eos_vlan_config = eos_vlan_grabber(eos_full_config)
    eos_config_paragraph = eos_paragraph_splitter(eos_config_file)
    eos_config_parameters = eos_config_extractor(eos_config_paragraph)
    voss_config = config_param(eos_config_parameters,voss_include_vlan_create,voss_isid,eos_vlan_config)
    with open('voss_config.txt', 'w') as w:
        w.write('\n'.join(voss_config))
        w.close()
    create_excel(eos_config_parameters)
elif eos_excel_confirm[0] == 'y':
    files_xlsx = []
    eos_vlan_config = {}
    #This obtains the current path
    current_path = os.getcwd()
    #This determines the file format to search for
    file_ext = r".xlsx"
    for file in os.listdir(current_path):
        if file.endswith(file_ext):
            print(os.path.join("", file))
            files_xlsx.append(os.path.join("", file))
    what_file = [inquirer.List('xlxs-files', message = "Which xlsx file below are you using as a template", choices = files_xlsx,)]
    xlsx_workbook_name = inquirer.prompt(what_file)
    xlsx_workbook = load_workbook(filename=xlsx_workbook_name['xlxs-files'])
    xlsx_sheets = xlsx_workbook.sheetnames
    what_worksheet = [inquirer.List('xlxs-sheets', message = "Which sheet will you be using for interface and VLAN configuration", choices = xlsx_sheets,)]
    xlsx_worksheet_name = inquirer.prompt(what_worksheet)
    eos_config_parameters = read_excel (xlsx_workbook,xlsx_worksheet_name['xlxs-sheets'])
    voss_config = config_param(eos_config_parameters,voss_include_vlan_create,voss_isid,eos_vlan_config)
    with open('voss_config.txt', 'w') as w:
        w.write('\n'.join(voss_config))
        w.close()
else:
    print ("You must answer 'y' to either using EOS backup .txt or 'y' to using the EXCEL template to create configuration")
if produce_excel[0] == 'y':
    print (read_excel)
    print ('An Excel document called configuration.xlxs has been created in the same directory as your script')
    